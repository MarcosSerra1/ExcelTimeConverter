# Ler dados da planilha 
import openpyxl

# Abrir planilha
planilha = openpyxl.load_workbook('CONVERSAO_HORAS_DECIMAIS.xlsx')
# Abrir pagina da planilha
conversao_sheet = planilha['conversao']

# Função para transformar horas em horas decimais
def horasDecimais(horas, minutos):
    minutos_decimal = minutos / 60
    minutos_formatados = float("{:.2f}".format(minutos_decimal)) # Formatando os numeros com duas casas decimais
    
    # Horas convertidas
    horas_decimal = horas + minutos_formatados
    return horas_decimal



# acessando linha de hora e minutos
for linha in conversao_sheet.iter_rows(min_row=2):
    horas = (linha[5].value) # Coluna das horas
    minutos = (linha[6].value) # Coluna dos minutos

    if horas is not None and minutos is not None:
        horas_decimal = horasDecimais(horas, minutos)
        # Adcionar as horas decimais de volta à planilha na coluna desejada (no caso 8)
        conversao_sheet.cell(row=1, column=8, value="Horas Decimais")  # Adicione o cabeçalho
        conversao_sheet.cell(row=linha[2].row, column=8, value=horas_decimal)

# Salvar a planilha com os valores calculados
planilha.save('CONVERSAO_HORAS_DECIMAIS_atualizado.xlsx')
print('Valores calculados salvos na planilha "CONVERSAO_HORAS_DECIMAIS_atualizado.xlsx"')
