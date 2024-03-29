from tkinter import filedialog, messagebox
from utils.converte_tempo_para_decimal import converte_tempo_para_decimal
import openpyxl


def calcular_horas_decimais():
    '''
    Realiza o cálculo das horas decimais a partir de um arquivo Excel.

    Esta função abre uma janela para permitir que o usuário selecione um arquivo Excel (.xlsx).
    Em seguida, carrega a planilha Excel selecionada, lê os dados da planilha 'conversao',
    calcula as horas decimais para cada linha que contém valores não nulos de horas e minutos,
    e atualiza a planilha 'evento_simplificado.xls' com os valores calculados.

    Se o usuário cancelar a seleção do arquivo Excel, a função não executa nenhuma operação.

    Args:
    - Nenhum.

    Returns:
    - Nenhum.

    Raises:
    - Nenhum.

    Exemplo de uso:
    - calcular_horas_decimais()
      Esta chamada inicia o processo de cálculo das horas decimais a partir de um arquivo Excel.
    '''

    # Abrindo uma janela para o usuário selecionar o arquivo Excel
    file_path = filedialog.askopenfilename(title="Selecione a planilha", filetypes=[("Planilha Excel", "*.xlsx")])

    # Verificando se o usuário cancelou a seleção do arquivo
    if not file_path:
        return

    # Carregando a planilha Excel
    planilha = openpyxl.load_workbook(file_path)
    conversao_sheet = planilha['conversao']  # Obtendo a planilha 'conversao'
    evento_simplificado_sheet = planilha['evento_simplificado.xls']  # Obtendo a planilha 'evento_simplificado.xls'


     # Define a linha inicial onde os dados serão colados na planilha de destino
    linha_destino = 2

    # Itera sobre as linhas da planilha de origem e copia os dados para a planilha de destino
    for linha in conversao_sheet.iter_rows(min_row=2):
        codigo = linha[0].value
        nome_funcionario = linha[1].value
        tipo_de_calculo = linha[2].value
        evento = linha[3].value
        valor_manual = linha[4].value

        # colocando os dados na outra planilha
        evento_simplificado_sheet.cell(row=linha_destino, column=1, value=codigo)
        evento_simplificado_sheet.cell(row=linha_destino, column=2, value=nome_funcionario)
        evento_simplificado_sheet.cell(row=linha_destino, column=3, value=tipo_de_calculo)
        evento_simplificado_sheet.cell(row=linha_destino, column=4, value=evento)
        evento_simplificado_sheet.cell(row=linha_destino, column=5, value=valor_manual)
        
        # incrementa a linha de destino para a próxima iteração
        linha_destino += 1


    # Iterando pelas linhas da planilha 'conversao'
    for linha in conversao_sheet.iter_rows(min_row=2):  # Começando da segunda linha
        horas = linha[5].value  # Obtendo o valor da coluna de horas
        minutos = linha[6].value  # Obtendo o valor da coluna de minutos

        # Verificando se horas e minutos não são nulos
        if horas is not None and minutos is not None:
            # Convertendo horas e minutos para o formato decimal
            hora_decimal = converte_tempo_para_decimal(horas=horas, minutos=minutos)
            # Atualizando a célula na planilha 'evento_simplificado.xls' com o valor decimal calculado
            evento_simplificado_sheet.cell(row=linha[2].row, column=6, value=hora_decimal)


    # Solicitando ao usuário o local para salvar a planilha modificada
    output_path = filedialog.asksaveasfilename(title="Salvar planilha", filetypes=[("Planilha Excel", "*.xlsx")])

    # Removendo a planilha 'conversao'
    planilha.remove(conversao_sheet)

    # Salvando a planilha modificada
    planilha.save(output_path)

    # Exibindo uma mensagem informando que o processo foi concluído
    messagebox.showinfo("Concluído", f"Valores calculados salvos em {output_path}")
