import tkinter as tk
from tkinter import filedialog, messagebox

import openpyxl


def horas_decimais(horas, minutos):
    minutos_decimal = minutos / 60
    minutos_formatados = float("{:.2f}".format(minutos_decimal))

    horas_decimal = horas + minutos_formatados
    return horas_decimal

def calcular_horas_decimais():
    file_path = filedialog.askopenfilename(title="Selecione a planilha", filetypes=[("Planilha Excel", "*.xlsx")])

    if not file_path:
        return

    planilha = openpyxl.load_workbook(file_path)
    conversao_sheet = planilha['conversao']
    evento_simplificado_sheet = planilha['evento_simplificado.xls']

    for linha in conversao_sheet.iter_rows(min_row=2):
        horas = linha[5].value
        minutos = linha[6].value

        if horas is not None and minutos is not None:
            horas_decimal = horas_decimais(horas, minutos)
            evento_simplificado_sheet.cell(row=linha[2].row, column=6, value=horas_decimal)

    planilha.remove(conversao_sheet)
    output_path = filedialog.asksaveasfilename(title="Salvar planilha", filetypes=[("Planilha Excel", "*.xlsx")])
    planilha.save(output_path)

    messagebox.showinfo("Concluído", f"Valores calculados salvos em {output_path}")

# Configuração da janela principal
root = tk.Tk()
root.title("Calculadora de Horas Decimais")
root.geometry("400x200")  # Definindo o tamanho da janela

# Botão para iniciar o cálculo
calcular_button = tk.Button(root, text="Calcular Horas Decimais", command=calcular_horas_decimais)
calcular_button.pack(pady=20)

# Botão para finalizar a aplicação
fechar_button = tk.Button(root, text="Fechar", command=root.destroy)
fechar_button.pack(pady=10)

# Iniciar o loop principal da interface gráfica
root.mainloop()
