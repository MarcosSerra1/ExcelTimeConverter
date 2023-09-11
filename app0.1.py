import openpyxl
import time
import os
import sys

# Obtém o diretório do executável, onde o script está sendo executado
exec_dir = os.path.dirname(sys.argv[0])

# Constrói o caminho completo para a planilha
planilha_path = os.path.join(exec_dir, 'CONVERSAO_HORAS_DECIMAIS_app1.xlsx')

# Verifique se o arquivo da planilha existe
if not os.path.exists(planilha_path):
    print('=' * 65)
    print(f'O arquivo {planilha_path} não foi encontrado.')
    print('=' * 65)
    time.sleep(5)
    sys.exit(1)

# Abrir a planilha
planilha = openpyxl.load_workbook(planilha_path)
conversao_sheet = planilha['conversao']

# Planilha de evento_simplificado.xls
eventoSimplificado = planilha['evento_simplificado.xls']

# Função para transformar horas em horas decimais
def horasDecimais(horas, minutos):
    minutos_decimal = minutos / 60
    minutos_formatados = float("{:.2f}".format(minutos_decimal)) # Formatando os numeros com duas casas decimais
    
    # Horas convertidas
    horas_decimal = horas + minutos_formatados
    return horas_decimal



# acessando linha de hora e minutos
for linha in conversao_sheet.iter_rows(min_row=2):
    horas = (linha[5].value) # Coluna das horas
    minutos = (linha[6].value) # Coluna dos minutos

    if horas is not None and minutos is not None:
        horas_decimal = horasDecimais(horas, minutos)
        # Adcionar as horas decimais de volta à planilha na coluna desejada (no caso na coluna 6 de nome Valor)
        eventoSimplificado.cell(row=linha[2].row, column=6, value=horas_decimal)

# Salvar a planilha com os valores calculados
planilha.remove(conversao_sheet) # Excluindo planilha conversão
planilha.save('evento_simplificado.xlsx')
print('=' * 65)
print('Valores calculados salvos na planilha "evento_simplificado.xlsx"')
print('=' * 65)
time.sleep(5)
